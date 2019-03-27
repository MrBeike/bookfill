# -*- coding:utf-8 -*-
import openpyxl as xl
import pandas as pd

# 打开excel统计表
def openbook(filename):
    xlsbook = xl.load_workbook(filename)
    xlsheets = xlsbook.get_sheet_names()
    sheet = xlsbook.get_sheet_by_name(xlsheets[0])
    return sheet

#读取excel表格数据
#dataSet:{row:[nume,deno]}   type:dict
def readData(sheet):
    dataSet = {}
    for i in range(5,70):  #sheet.max_row  根据实际情况调整
        nume = sheet.cell(i, 12).value  #L列  分子
        deno = sheet.cell(i, 14).value  #N列  分母
        data = {i:[nume,deno]}
        dataSet.update(data)
    return dataSet


filename = r"附件：附件二：中国普惠金融指标体系（2018年版）填报模板.xlsx"
s = openbook(filename)
k = readData(s)


# case1:sum
# 利用DataFrame进行汇总操作
kk = pd.DataFrame(k,index=["L","N"])
kkk = kk*3
print(kkk)
print(kkk.loc["L"][5])

DataFrame.bookfill(row,dataFrame)

sheet.cell(i,12).value = kkk.loc["L"][i]
sheet.cell(i,14).value = kkk.loc["N"][i]
# case2:seperate
sum1= readData()   #修改后的汇总文件
sum2 = sum(readData() *3) #汇总分文件
rate = compare(sum1,sum2)
fixData(book,rate)
