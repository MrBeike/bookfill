# -*- coding:utf-8 -*-
import openpyxl as xl
import numpy as np
import pandas as pd


class Book:
    def __init__(self):
        self.filenames = self.readIni()
        # 初始化空DataFrame用于后期汇总（有没有更简单的办法？）
        series = np.zeros((2, 65))
        self.dataSum = pd.DataFrame(series, index=["L", "N"], columns=range(5, 70))
        self.datacheck = pd.DataFrame(series, index=["L", "N"], columns=range(5, 70))

    # 读取配置文件
    # filenames:[filename1,filename2...] type:list
    def readIni(self):
        filenames = []
        file = open("ini.txt", "r")
        lines = file.readlines()
        # 获取每一行的文件名，设定最后一行为汇总表格
        for line in lines:
            line = line.strip('\n')
            filenames.append(line)
        return filenames

    # 打开一个文件，并返回这个文件openpyxl对象和第一个sheet
    def openBook(self, filename):
        xlsbook = xl.load_workbook(filename)
        xlsheets = xlsbook.get_sheet_names()
        sheet = xlsbook.get_sheet_by_name(xlsheets[0])
        return xlsbook, sheet

    # 读取excel表格数据
    # dataSet:{row:[nume,deno]}   type:dict
    # dataSets:[[dataSet1],[dataSet2]...]  type:list
    def readData(self, sheet):
        dataSet = {}
        for i in range(5, 70):  # sheet.max_row  根据实际情况调整
            nume = sheet.cell(i, 12).value  # L列  分子
            deno = sheet.cell(i, 14).value  # N列  分母
            data = {i: [nume, deno]}
            dataSet.update(data)
        return dataSet

    # 将数据填写到指定表格
    # @param:sheet
    # @param:data pd.DataFrame.object
    def bookfill(self, sheet, data):
        for row in data.columns.values.tolist():
            try:
                sheet.cell(row, 12).value = data.loc["L"][row]
                sheet.cell(row, 14).value = data.loc["N"][row]
            except AttributeError:
                continue
        return

    # 检查分表与汇总表数据是否一致
    def dataCheck(self, filenames):
        for filename in filenames[:-1]:
            book, sheet = self.openBook(filename)
            dataSet = self.readData(sheet)
            dataFrame = pd.DataFrame(dataSet, index=["L", "N"])
            self.datacheck += dataFrame
        sumBook, sumSheet = self.openBook(filenames[-1])
        sumDataSet = self.readData(sumSheet)
        sumDataFram = pd.DataFrame(sumDataSet, index=["L", "N"])
        check = sumDataFram.sub(self.datacheck)
        check = check.fillna(0.0)
        flag = check.any(axis=None)
        flags = []
        if flag:
            flagdict = check.any(axis=0).to_dict()
            for key, value in flagdict.items():
                if value:
                    flags.append(key)
                else:
                    continue
            print("以下行数据存在问题：", flags)
        else:
            print("数据完整性校验通过")


    # Case 1
    # 汇总所有分表的数据
    # data:pd.DataFrame({5:XX,6:xx,....})  type:pd.DataFrame.object
    # 是否转换回dict?
    def bookSum(self, filenames):
        for filename in filenames[:-1]:
            book, sheet = self.openBook(filename)
            dataSet = self.readData(sheet)
            dataFrame = pd.DataFrame(dataSet, index=["L", "N"])
            self.dataSum += dataFrame
        sumBook, sumSheet = self.openBook(filenames[-1])
        self.bookfill(sumSheet, self.dataSum)
        sumBook.save(filenames[-1])
        print("表格汇总完成，将进行数据完整性检查")
        return

    # Case 2
    # 对比汇总表与分表汇总数据差别，等比例修改分表数据
    def bookFix(self, filenames):
        for filename in filenames[:-1]:
            book, sheet = self.openBook(filename)
            dataSet = self.readData(sheet)
            dataFrame = pd.DataFrame(dataSet, index=["L", "N"])
            self.dataSum += dataFrame
        sumBook, sumSheet = self.openBook(filenames[-1])
        sumDataSet = self.readData(sumSheet)
        sumDataFram = pd.DataFrame(sumDataSet, index=["L", "N"])
        rate = sumDataFram.div(self.dataSum)
        for filename in filenames[:-1]:
            book, sheet = self.openBook(filename)
            dataSet = self.readData(sheet)
            dataFrame = pd.DataFrame(dataSet, index=["L", "N"])
            dataFrameFix = dataFrame.mul(rate)
            self.bookfill(sheet, dataFrameFix)
            book.save(filename)
        print("表格修改完成，将进行数据完整性检查")
        return


if __name__ == '__main__':
    print("-----------------------------------------------------------------------------")
    print("----------开始之前请确保已完成ini.txt文件存在、文件名正确、对应文件存在----------")
    print("1.汇总分地区表")
    print("2.修改分地区表")
    print("3.汇总校验")
    book = Book()
    c = eval(input("请选择操作项:"))
    if c == 1:
        book.bookSum(book.filenames)
        book.dataCheck(book.filenames)
    elif c == 2:
        book.bookFix(book.filenames)
        book.dataCheck(book.filenames)
    elif c == 3:
        book.dataCheck(book.filenames)
    else:
        print("嘿，大兄弟。你这干啥呢？")