# -*- coding:utf-8 -*-
import openpyxl as xl

#BookFill(copy,word)
#@param copy 问卷份数
#@param word 空白格所用标识，默认为字符'n'.
class BookFill:
    def __init__(self, copy,word = 'n'):
        # 确定表格映射关系
        # 1.金融产品表（答案从第7位开始，注意第9位为空时，使用N？）
        infoRelation = [[], [], [], [], [], [17, 18]]
        productRelation = [[5, 6, 7, 8], [15, 16], [23, 24], [31], [38, 39, 40], [47, 48, 49], [56, 57]]
        knowledgeRelation = [[6, 7, 8], [10, 11, 12], [14, 15, 16, 17], [19, 20, 21], [23, 24], [26, 27, 28, 29, 30],
                             [32, 33, 34], [36, 37, 38], [40, 41, 42, 43], [45, 46, 47]]
        actionRelation = [[6, 7, 8, 9], [11, 12, 13], [15, 16, 17, 18, 19], [21, 22, 23, 24], [26, 27, 28, 29],
                          [31, 32, 33, 34], [36, 37, 38], [40, 41, 42], [44, 45, 46, 47], [49, 50, 51, 52, 53]]
        self.relation = (infoRelation, productRelation, knowledgeRelation, actionRelation)
        self.filename = r"附件：附件五：普惠金融问卷调查指标录入和计算模板.xlsx"
        self.copy = copy
        self.colum = 4 + self.copy
        self.word = word
        self.sheets = self.openbook()
        self.answer = self.readdata()

    # 打开excel统计表
    def openbook(self):
        self.xlsbook = xl.load_workbook(self.filename)
        xlsheets = self.xlsbook.get_sheet_names()
        infoSheet = self.xlsbook.get_sheet_by_name(xlsheets[0])
        productSheet = self.xlsbook.get_sheet_by_name(xlsheets[1])
        knowledgeSheet = self.xlsbook.get_sheet_by_name(xlsheets[2])
        actionSheet = self.xlsbook.get_sheet_by_name(xlsheets[3])
        sheets = (infoSheet, productSheet, knowledgeSheet, actionSheet)
        return sheets

    # 数据合法性检查
    def datacheck(self):
        file = open("data.txt", "r")
        lines = file.readlines()
        flag = []
        for i in range(len(lines)):
            line = lines[i].strip('\n')
            if len(line) == 33:
                continue
            else:
                flag.append(i + 1)
        if len(flag) != 0:
            print("以下行数据存在错误：", flag)
            # 制成exe文件时需要调整跳出方法
            s = input("请关闭程序并检查data.txt文件")
            return False
        else:
            return True

    # 读取一份调查问卷统计数据
    def readdata(self):
        file = open("data.txt", "r")
        lines = file.readlines()
        # for line in lines:
        line = lines[self.copy]
        line = line[:33]
        # 获取一条统计序列数，然后拆分
        for i in range(len(line)):
            infoAnswer = line[0:6]
            productAnswer = line[6:13]
            knowledgeAnswer = line[13:23]
            actionAnswer = line[23:33]
        answer = (infoAnswer, productAnswer, knowledgeAnswer, actionAnswer)
        return answer

    # 接收答案和表格对应关系，输出单元格坐标
    def position(self, answer, relation):
        positions = []
        # flag标记题目为选择题（0）还是填空题（1）或是不填数值（-1）
        for i in range(len(answer)):
            flag = 0
            key = answer[i]
            # 判断该题目答案是否为空
            if key == self.word:
                flag = -1
                key, row = 0, 0
                position = [flag, key, row]
            else:
                key = eval(key)
                try:
                    row = relation[i][key - 1]
                # 如果数组角标超过，则表示是填空题，取唯一位置（即第一个位置）
                except IndexError:
                    try:
                        row = relation[i][0]
                        flag = 1
                    # 如果映射为空，则位置数组为空，表示不填数值，raise indexError
                    except IndexError:
                        row = 0
                        flag = -1
                position = [flag, key, row]
            # print(position)
            positions.append(position)
        return positions

    # 接受列数（实际为问卷份数）
    def fill(self, sheet, positions):
        for position in positions:
            if position[0] == 0:
                sheet.cell(position[2], self.colum).value = 1
            elif position[0] == 1:
                sheet.cell(position[2], self.colum).value = position[1]
            else:
                continue
        self.xlsbook.save(self.filename)
        print('数据读取写入完成:', self.copy + 1, sheet)

    def screenpause(self):
        for i in range(1000, 0):
            print('程序将在' + i + '秒后自动关闭')


#程序入口
if __name__ == '__main__':
    c = eval(input("请输入问卷份数:"))
    for k in range(c):
        bookfill = BookFill(k)
        flag = bookfill.datacheck()
        if flag:
            sheets = bookfill.sheets
            answer = bookfill.answer
            for i in range(0, 4):
                positions = bookfill.position(answer[i], bookfill.relation[i])
                bookfill.fill(sheets[i], positions)
